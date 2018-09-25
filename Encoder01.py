#! /usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
wb_rui=openpyxl.load_workbook('流衣表情差分.xlsx')
sheet_rui=wb_rui['シート1']

import re

class SheetData:
    wb_rui = openpyxl.load_workbook('流衣表情差分.xlsx')
    wb_jin = openpyxl.load_workbook('仁表情差分.xlsx')
    wb_tobari = openpyxl.load_workbook('帷表情差分.xlsx')
    wb_saki = openpyxl.load_workbook('早希表情差分.xlsx')
    wb_mei = openpyxl.load_workbook('命表情差分.xlsx')
    wb_miu = openpyxl.load_workbook('美羽表情差分.xlsx')
    wb_rumi = openpyxl.load_workbook('瑠美表情差分.xlsx')
    wb_sayo1_3 = openpyxl.load_workbook('小夜表情差分(1~3).xlsx')
    wb_shiori = openpyxl.load_workbook('栞表情差分.xlsx')
    wb_ciel = openpyxl.load_workbook('シエル青目差分.xlsx')
    sheet_ciel = wb_ciel['シート1']
    sheet_shiori = wb_shiori['シート1']
    sheet_rui = wb_rui['シート1']
    sheet_jin = wb_jin['シート1']
    sheet_tobari = wb_tobari['シート1']
    sheet_saki = wb_saki['シート1']
    sheet_mei = wb_mei['シート1']
    sheet_miu = wb_miu['シート1']
    sheet_rumi = wb_rumi['シート1']
    sheet_sayo = wb_sayo1_3['シート1']

    def loadPartsNum(self, face_num, name):
        partsList = []
        for index, cell_obj in enumerate(list(getattr(self, 'sheet_' + name).rows)[face_num][1:]):
            val=cell_obj.value
            if re.match(r'[0-9]+',val):
                partsList.append(re.match(r'[0-9]+',val).group())
        return partsList
    
    def loadCharaSize(self, name):
        return list(getattr(self, 'sheet_' + name).rows)[0][0].value

class CreateCharaTag:
    def __init__(self):
        self.__name = ""
        self.__last_name = ""
        self.__zoom = ""
        self.__sheetData = SheetData()

    def getName(self):
        return self.__name
    def setName(self, name):
        self.__name = name
    def getLastName(self):
        return self.__last_name
    def setLastName(self, name):
        self.__last_name = name
    def getZoom(self):
        return self.__zoom
    def getSheetData(self):
        return self.__sheetData

    def processBrackets(self, brackets):
        self.__name = re.search(r'[a-z]+', brackets).group()
        self.__zoom = brackets[-3:-1].lower()
        face_code = re.search(u'[0-9]{1}-+[0-9]+', brackets).group()
        scene_num = face_code[:1]
        face_num = face_code[2:]

        partsList = self.__sheetData.loadPartsNum(int(face_num), self.__name)
        tag = '[chara_part name=' + self.__name + ' body=' + scene_num + ' eye=' + partsList[0] + ' mouth=' + partsList[1] + ' eyeblow=' + partsList[2] + ' effect=' + partsList[3] + ']'

        return tag

def processTalkBracket(brackets):
    pos_open_bracket = re.search('「|《', brackets).start()
    name = '#' + re.sub(r'　+', "", brackets[:pos_open_bracket]) + '\n'
    talk = re.sub(r'「{1}|」{1}|《{1}|》{1}', "", brackets[pos_open_bracket:])
    return name + talk

def processLineWithoutBracket(line):
    return re.sub(r'　+|」{1}', "", line)

def processBGFOIBracket(brackets):
    pos_slash = brackets.find("/")
    return '[bg_foi storage="' + brackets[1:pos_slash] + '.jpg"]'

class CreateMusicTag:    
    wb_music = openpyxl.load_workbook('0_BGMリスト.xlsx')
    wb_se = openpyxl.load_workbook('0_SEリスト.xlsx')
    sheet_music = wb_music['シート1']
    sheet_se = wb_se['シート1']

    def processBgMusicFOITag(self, sound_num, bg_name):
        return '[bg_music_foi bgstr="' + bg_name + '.jpg" musicstr="' + sound_num + '.' + list(self.sheet_music.rows)[int(sound_num)][1].value + '.ogg"]'

    def processBgFOIMusicFOTag(self, bg_name):
        return '[bg_foi_music_fo bgstr="' + bg_name + '.jpg"]'

    def processSoundTag(self, sound_num, sound_type, sound_sub_type):
        if sound_type in ['BGM']:
            sound_name = list(self.sheet_music.rows)[int(sound_num)][1].value
            if sound_sub_type == 'CO':
                return '[stopbgm]'
            elif sound_sub_type == 'FO':
                return '[fadeoutbgm]'
            elif sound_sub_type == 'CI':
                return '[playbgm storage="' + sound_name + '.ogg" loop=true]'
            elif sound_sub_type == 'FI':
                return '[fadeinbgm storage="' + sound_name + '.ogg" loop=true]'
        elif sound_type in ['BGS', 'SE']:
            sound_name = list(self.sheet_se.rows)[int(sound_num)][1].value  
            if sound_sub_type == 'CO':
                return '[stopse]'
            elif sound_sub_type == 'FO':
                return '[fadeoutse]'
            elif sound_type == 'BGS':
                if sound_sub_type == 'CI':
                    return '[playse storage="' + sound_name + '.ogg" loop=true]'
                elif sound_sub_type == 'FI':
                    return '[fadeinse storage="' + sound_name + '.ogg" loop=true]'
            elif sound_type == 'SE':
                if sound_sub_type == 'CI':
                    return '[playse storage="' + sound_name + '.ogg" loop=false]'
                elif sound_sub_type == 'FI':
                    return '[fadeinse storage="' + sound_name + '.ogg" loop=false]'

class ConversionBracketsToTags:
    def __init__(self):
        self.__createCharaTags = [CreateCharaTag(), CreateCharaTag()]
        self.__createMusicTag = CreateMusicTag()
        self.__tags = []
        self.__flag_bg_foi = False
        self.__flag_chara_hide_all = False
        self.__bg_foi_name = ""
    def conversion(self, array_of_brackets):
        self.__tags = ["", ""]
        if len(array_of_brackets) < 2:
            self.__createCharaTags[1].setName("")
        for index, brackets in enumerate(array_of_brackets):
            if brackets.find('「') != -1:
                self.__tags[index] += processTalkBracket(brackets)

            if brackets.find('【') != -1:
                if brackets.find('表示なし') != -1 :
                    if self.__flag_chara_hide_all == False:
                        self.__tags[0] += '[chara_hide_all]'
                        self.__flag_chara_hide_all = True
                else:
                    self.__flag_chara_hide_all = False
                    self.__tags[index] += self.__createCharaTags[index].processBrackets(brackets)
                    if index > 0:
                        self.__tags[0] += '\n'
                    if self.__createCharaTags[0].getLastName() not in ['', self.__createCharaTags[0].getName(), self.__createCharaTags[1].getName()]:
                        self.__tags[index] += '\n[chara_hide name="' + self.__createCharaTags[0].getLastName() + ']'
                    if self.__createCharaTags[1].getLastName() not in ['', self.__createCharaTags[0].getName(), self.__createCharaTags[1].getName()]:   
                        self.__tags[index] += '\n[chara_hide name="' + self.__createCharaTags[1].getLastName() + ']' 
                    if self.__createCharaTags[index].getName() not in [self.__createCharaTags[0].getLastName(), self.__createCharaTags[1].getLastName()]:
                        if index > 0:
                            chara_size = self.__createCharaTags[index].getSheetData().loadCharaSize(self.__createCharaTags[1].getName())
                            self.__tags[index] += '\n[chara_add name="' + self.__createCharaTags[1].getName() + '" org="' + self.__createCharaTags[0].getName() + '" zoom=' + self.__createCharaTags[0].getZoom() + ' size=' + chara_size + ']'
                        else:
                            self.__tags[index] += '\n[s_show name="' + self.__createCharaTags[index].getName() + '" zoom=' + self.__createCharaTags[0].getZoom() + ']'

            if brackets.find('〈') != -1:
                if brackets.find('FOI') != -1:
                    self.__tags[index] += processBGFOIBracket(brackets)
                    pos_slash = brackets.find("/")
                    self.__bg_foi_name = brackets[1:pos_slash]
                    if len(array_of_brackets) >= 2:
                        self.__flag_bg_foi = True
                    else:
                        self.__flag_bg_foi = False
                else:
                    self.__tags[index] += brackets    

            if brackets.find('〔') != -1:
                sound_type = brackets[1:brackets.find(':')]
                sound_sub_type = brackets[brackets.find('/') + 1:-1]
                sound_num = re.search('[0-9]+', brackets).group()
                if self.__flag_bg_foi:             
                    if brackets.find('FOI') != -1:
                        self.__tags[0] = self.__createMusicTag.processBgMusicFOITag(sound_num, self.__bg_foi_name)
                        self.__tags[1] = ""
                    elif brackets.find('FO') != -1:
                        self.__tags[0] = self.__createMusicTag.processBgFOIMusicFOTag(self.__bg_foi_name)
                        self.__tags[1] = "" 
                    else:
                        self.__tags[0] = '〈' + self.__bg_foi_name + '/FOI〉'
                        self.__tags[1] = brackets                      
                    self.__flag_bg_foi = False
                else:                    
                    self.__tags[index] += self.__createMusicTag.processSoundTag(sound_num, sound_type, sound_sub_type)

            if index + 1 >= len(array_of_brackets):
                self.__tags[index] += '\n'

        self.__createCharaTags[0].setLastName(self.__createCharaTags[0].getName())
        self.__createCharaTags[1].setLastName(self.__createCharaTags[1].getName())
        return self.__tags[0] + self.__tags[1]

convBtoT = ConversionBracketsToTags()
file_name = input('filename>')
flag_empty_sharp = False
line_num = 0
with open(file_name) as f:
    text = []
    while True:
        s_line = f.readline()
        line_num += 1
        print(line_num)
        if not s_line:
            break

        brackets = re.findall('「{1}.*?」{1}|《{1}.*?》{1}|【{1}.*?】{1}|〔{1}.*?〕{1}|〈{1}.*?〉{1}', s_line)
        if not brackets:
            if re.search('「|《', s_line):
                text.append(processTalkBracket(s_line))
                flag_empty_sharp = False
            else:
                if flag_empty_sharp:
                    text.append('\n#')
                    flag_empty_sharp = False
                text.append(processLineWithoutBracket(s_line))
        elif re.search('「|《', s_line):
            if re.search('」|》', s_line):
                flag_empty_sharp = True
                if len(brackets) < 2:
                    text.append(processTalkBracket(s_line))
                else:
                    temp_line = re.sub('%s'%brackets[1], "", s_line)  
                    text.append(processTalkBracket(temp_line))
                    brackets.pop(0)
                    text.append(convBtoT.conversion(brackets))                       
            else:
                temp_line = re.sub('%s'%brackets, "", s_line)  
                text.append(processTalkBracket(temp_line))
                text.append(convBtoT.conversion(brackets))      
        else:
            text.append(convBtoT.conversion(brackets))

with open("result01.txt", "w", encoding="utf-8") as f:
    f.writelines(text)

